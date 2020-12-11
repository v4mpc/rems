from openpyxl import load_workbook
import logging
import datetime
from django.conf import settings
import os
from pathlib import Path
from openpyxl.drawing.image import Image


class WorkBook:
    def __init__(self, sheet_name):
        self.module_dir = os.path.dirname(__file__)
        self.sample_arf_path = os.path.join(
            self.module_dir, 'static/rems_api/arf.xlsx')  # "arf.xlsx"  # erf.xlsx
        self.wb = load_workbook(filename=self.sample_arf_path)
        self.sheet_name = sheet_name  # "Advance Request"  # Tanzania Expense Report
        self.sheet = self.wb[self.sheet_name]

    def write_cell(self, cell_address, cell_value):
        self.sheet[cell_address] = cell_value

    def get_number_of_nights(self, start_date, end_date):
        # date dd-mm-yyy
        start_date = datetime.datetime.strptime(start_date, "%d-%m-%Y")
        end_date = datetime.datetime.strptime(end_date, "%d-%m-%Y")
        delta_date = end_date-start_date
        return delta_date.days

    def save(self):
        self.wb.save(self.generate_file_name())

    def read_cell(self, cell_address):
        return self.sheet[cell_address].value

    def exists(self):
        return Path(self.generate_file_name()).is_file()

    def delete(self, file_name=None):
        os.remove(self.generate_file_name(file_name))

    def init(self, validated_data):
        self.region = validated_data['location'].name
        self.date_of_request = validated_data['date_of_request']
        self.user = validated_data['user']
        self.name = self.user.profile.name
        self.address = validated_data['address']
        self.purpose = validated_data['purpose']
        self.period_of_travel_from = validated_data['start_date']
        self.period_of_travel_to = validated_data['end_date']
        self.signature = ''
        self.mes = validated_data['mes']
        self.lodgings = validated_data['lodgings']
        self.other_costs = validated_data['other_costs']
        self.signature_date = self.date_of_request
        self.fields = {
            'date_of_request': 'G4',
            'name': 'B8',
            'address': 'B9',
            'purpose': 'B13',
            'period_of_travel_from': 'B16',
            'period_of_travel_to': 'E16',
            'me_destination': ['B27', 'B28', 'B29', 'B30', 'B31', 'B21'],
            'me_no_of_days': ['C27', 'C28', 'C29', 'C30', 'C31', 'C21'],
            'me_amount': ['D27', 'D28', 'D29', 'D30', 'D31', 'D21'],
            'me_rate': ['E27', 'E28', 'E29', 'E30', 'E31', 'E21'],
            'lodge_destination': ['B41', 'B42', 'B43', 'B44', 'B45', 'B46'],
            'lodge_no_of_nights': ['C41', 'C42', 'C43', 'C44', 'C45', 'C46'],
            'lodge_amount': ['D41', 'D42', 'D43', 'D44', 'D45', 'D46'],
            'lodge_rate': ['E41', 'E42', 'E43', 'E44', 'E45', 'E46'],
            'other_purpose': ['B55', 'B56', 'B57', 'B58', 'B59'],
            'other_amount': ['D55', 'D56', 'D57', 'D58', 'D59'],
            'signature': 'B65',
            'signature_date': 'G65',
            'approved_by': 'B68',
            'approve_date': 'G68'
        }

    def generate_file_name(self, file_name=None):
        # region = self.region
        file_name = file_name if file_name is not None else self.get_file_name()
        return os.path.join(
            self.module_dir, f'static/rems_api/{file_name}')

    def get_file_name(self):
        region = "_".join(self.region.split(' '))
        sheet_name = "_".join(self.sheet_name.split(' '))
        name = "_".join(self.name.split(' '))
        date_travel = self.date_of_request
        return f"{name}_{sheet_name}_{region}_{date_travel}.xlsx"

    def write_field_data(self, field, data):
        field_data = self.fields[field]
        if isinstance(field_data, list):

            # we have a list, data should be a list
            x = 0
            while len(data) > x:
                cell_address = self.fields[field][x]
                cell_value = data[x]
                if (field == 'lodge_rate') or (field == 'me_rate'):
                    self.write_cell(cell_address, f'{cell_value}%')
                else:
                    self.write_cell(cell_address, cell_value)

                x += 1

        else:
            # we have a string, data is a string
            cell_address = self.fields[field]
            cell_value = data
            self.write_cell(cell_address, cell_value)

    def write_and_save(self):
        self.write_field_data('date_of_request', self.date_of_request)
        self.write_field_data('name', self.name)
        # self.write_field_data('address', self.address)
        self.write_field_data('purpose', self.purpose)
        self.write_field_data(
            'period_of_travel_from', self.period_of_travel_from)
        self.write_field_data('period_of_travel_to',
                              self.period_of_travel_to)
        # me
        me = self.transform_for_write(self.mes)
        self.write_field_data('me_destination', me['destination'])
        self.write_field_data('me_no_of_days', me['no_of_nights'])
        self.write_field_data('me_amount', me['daily_rate'])
        self.write_field_data('me_rate', me['percentage_of_daily_rate'])
        # lodging
        lodging = self.transform_for_write(self.lodgings)
        self.write_field_data('lodge_destination', lodging['destination'])
        self.write_field_data('lodge_no_of_nights', lodging['no_of_nights'])
        self.write_field_data('lodge_amount', lodging['daily_rate'])
        self.write_field_data(
            'lodge_rate', lodging['percentage_of_daily_rate'])
        # other costs
        other_cost = self.transform_for_write(self.other_costs)
        if "purpose" in other_cost and "amount" in other_cost:
            self.write_field_data('other_purpose', other_cost['purpose'])
            self.write_field_data('other_amount', other_cost['amount'])

        self.write_field_data('signature_date', self.signature_date)
        self.save()

    def sign(self, file_name):
        image_path = os.path.join(
            self.module_dir, 'static/rems_api')
        image_path = os.path.join(image_path, file_name)
        signature = Image(image_path)
        signature.height = 54
        signature.widhth = 96
        self.sheet.add_image(signature)
        self.save()

    def transform_for_write(self, list_of_costs):
        dict_of_costs = {}
        for dict_of_cost in list_of_costs:
            for key, value in dict_of_cost.items():
                if key == 'start_date' or key == 'end_date':
                    continue
                try:
                    dict_of_costs[key].append(value)
                except KeyError:
                    dict_of_costs[key] = []
                    dict_of_costs[key].append(value)

        return dict_of_costs


class Erf(WorkBook):
    def __init__(self, file_to_edit):
        # super(WorkBook, self).__init__('Tanzania Expense Report')
        self.module_dir = os.path.dirname(__file__)
        self.file_to_edit = file_to_edit
        self.sample_arf_path = os.path.join(
            self.module_dir, f'static/rems_api/'+file_to_edit)  # "arf.xlsx"  # erf.xlsx
        self.wb = load_workbook(filename=self.sample_arf_path)
        # "Advance Request"  # Tanzania Expense Report
        self.sheet_name = 'Tanzania Expense Report'
        self.sheet = self.wb[self.sheet_name]

    def save(self):
        self.wb.save(self.generate_file_name(self.file_to_edit))

    def init(self, validated_data, mes_data, lodgings_data, other_costs_data):
        self.region = validated_data['location'].name
        self.date_of_request = validated_data['date_of_request']
        self.user = validated_data['user']
        self.name = self.user.profile.name
        self.address = validated_data['address']
        self.purpose = validated_data['purpose']
        self.period_of_travel_from = validated_data['start_date']
        self.period_of_travel_to = validated_data['end_date']
        self.signature = ''
        self.mes = mes_data
        self.lodgings = lodgings_data
        self.logding_max = validated_data['location'].lodging
        self.other_costs = other_costs_data
        self.signature_date = self.date_of_request
        self.fields = {
            'name': 'D3',
            'address': 'D4',
            'purpose': 'C14',
            'me_date': ['A26', 'A27', 'A28', 'A29', 'A30', 'A31', 'A32'],
            'me_destination': ['I26', 'I27', 'I28', 'I29', 'I30', 'I31', 'I32'],
            'me_no_of_days': ['M26', 'M27', 'M28', 'M29', 'M30', 'M31', 'M32'],
            'me_amount':  ['L26', 'L27', 'L28', 'L29', 'L30', 'L31', 'L32'],

            'lodge_date': ['A38', 'A39', 'A40', 'A41', 'A42', 'A43', 'A44'],
            'lodge_destination': ['C38', 'C39', 'C40', 'C41', 'C42', 'C43', 'C44'],
            'lodge_max': ['H38', 'H39', 'H40', 'H41', 'H42', 'H43', 'H44'],
            'lodge_no_of_nights': ['K38', 'K39', 'K40', 'K41', 'K42', 'K43', 'K44'],
            'lodge_actual_cost': ['J38', 'J39', 'J40', 'J41', 'J42', 'J43', 'J44'],

            'other_date': ['A68', 'A69'],
            'other_purpose': ['C68', 'C69'],
            'other_cost': ['G68', 'G69'],
            'other_receipt_no': ['I68', 'I69'],


            'signature_of_traveler': 'C18',
            'traveler_signature_date': 'M18',
            'confirmed_by': 'C19',
            'confirmed_date': 'M19'
        }

    def write_and_save(self):

        self.write_field_data('purpose', self.purpose)
        me = self.transform_for_write(self.mes)
        self.write_field_data('me_date', me['date'])
        self.write_field_data('me_destination', me['destination'])
        self.write_field_data('me_no_of_days', me['no_of_nights'])
        self.write_field_data('me_amount', me['daily_rate'])

        lodging = self.transform_for_write(self.lodgings)
        self.write_field_data('lodge_date', lodging['date'])
        self.write_field_data('lodge_destination', lodging['destination'])
        self.write_field_data('lodge_no_of_nights', lodging['no_of_nights'])
        self.write_field_data('lodge_actual_cost', lodging['daily_rate'])
        self.write_field_data(
            'lodge_max', [self.logding_max]*len(self.lodgings))

        other_cost = self.transform_for_write(self.other_costs)
        if "purpose" in other_cost and "cost" in other_cost:
            self.write_field_data('other_purpose', other_cost['purpose'])
            self.write_field_data('other_cost', other_cost['cost'])
            self.write_field_data('other_receipt_no', other_cost['receipt_no'])

        self.save()


if __name__ == "__main__":
    arf = WorkBook()
    print(arf.get_number_of_nights('21-09-2020', '27-09-2020'))
